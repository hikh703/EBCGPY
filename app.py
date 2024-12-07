from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import io
import os
import zipfile
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
import barcode
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from flask import send_from_directory

app = Flask(__name__, static_folder="static")
CORS(app)

@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")

@app.route("/generate-labels", methods=["POST", "OPTIONS"])
def generate_labels():
    if request.method == "OPTIONS":
        # Handle preflight request
        return jsonify({"message": "Preflight check"}), 200

    try:
        # Step 1: Retrieve uploaded file and user number
        uploaded_file = request.files.get("file")
        user_number = request.form.get("user_number")
        
        if not uploaded_file or not user_number:
            return jsonify({"error": "Fichier et numéro de portant requis."}), 400

        # Step 2: Open Excel file using openpyxl
        workbook = load_workbook(uploaded_file)
        sheet = workbook.active

        # Ensure required columns exist
        headers = {cell.value: idx for idx, cell in enumerate(sheet[1])}
        required_columns = {"Nom", "Prix", "Valeur Option1"}
        missing_columns = required_columns - headers.keys()
        if missing_columns:
            return jsonify({"error": f"Colonnes manquantes: {missing_columns}"}), 400

        # Add "Code-barres" column if not present
        if "Code-barres" not in headers:
            headers["Code-barres"] = len(headers)
            sheet.cell(row=1, column=headers["Code-barres"] + 1).value = "Code-barres"

        # Generate barcodes and labels
        today = datetime.now().strftime("%d%m%y")
        dpi = 600  # Change DPI to 600
        label_width = int(60 * dpi / 25.4)  # 60mm to pixels at 600 DPI
        label_height = int(30 * dpi / 25.4)  # 30mm to pixels at 600 DPI

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w") as zf:
            for idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=1):
                # Retrieve row data
                nom = row[headers["Nom"]].value
                prix = row[headers["Prix"]].value
                valeur_option1 = row[headers["Valeur Option1"]].value

                # Generate barcode
                barcode_code = f"SELEC{today}{user_number}{str(idx).zfill(4)}"
                row[headers["Code-barres"]].value = barcode_code

                # Create label image
                label = Image.new("RGB", (label_width, label_height), "white")
                draw = ImageDraw.Draw(label)

                # Load custom font (optional)
                try:
                    font_path = "arial.ttf"  # Replace with your font path
                    font_large = ImageFont.truetype(font_path, 90)  # Increased font size for better visibility
                    font_small = ImageFont.truetype(font_path, 80)  # Increased font size for better visibility
                except IOError:
                    font_large = font_small = ImageFont.load_default()

                # Prepare label content
                text_elements = []
                if nom:
                    text_elements.append((nom, font_large))
                text_elements.append((f"Taille: {valeur_option1 if valeur_option1 else ''}", font_small))
                if prix:
                    text_elements.append((f"{prix}€", font_small))

                # Calculate total text height
                total_text_height = sum(
                    draw.textbbox((0, 0), text, font=font)[3] - draw.textbbox((0, 0), text, font=font)[1] + 10
                    for text, font in text_elements
                )

                # Add extra space between "Taille" and "Prix"
                total_text_height += 40  # Additional space between Taille and Prix

                # Calculate text and barcode positions
                remaining_space = label_height - (int(label_height * 0.4) + 10)  # Space for barcode + spacing
                text_y_start = (remaining_space - total_text_height) // 2

                # Draw text
                for i, (text, font) in enumerate(text_elements):
                    text_width = draw.textbbox((0, 0), text, font=font)[2]
                    draw.text(
                        ((label_width - text_width) // 2, text_y_start),
                        text,
                        fill="black",
                        font=font,
                    )
                    text_height = draw.textbbox((0, 0), text, font=font)[3] - draw.textbbox((0, 0), text, font=font)[1]
                    text_y_start += text_height + (30 if i == 1 else 35)  # Extra space after Taille (i == 1)

                # Generate barcode
                EAN = barcode.get_barcode_class("code128")
                ean = EAN(barcode_code, writer=ImageWriter())
                barcode_image = ean.render(writer_options={"module_height": 10, "dpi": 600})  # Set barcode DPI to 600

                # Resize and paste barcode
                barcode_width = int(label_width * 1)  # 100% width of the label
                barcode_height = int(label_height * 0.5)  # 50% height of the label
                barcode_image = barcode_image.resize((barcode_width, barcode_height), Image.LANCZOS)
                barcode_x = (label_width - barcode_width) // 2
                barcode_y = label_height - barcode_height
                label.paste(barcode_image, (barcode_x, barcode_y))

                # Save label as PNG in memory
                label_buffer = io.BytesIO()
                label.save(label_buffer, format="PNG")
                label_buffer.seek(0)

                # Add to ZIP
                zf.writestr(f"{barcode_code}_label.png", label_buffer.read())

            # Add updated Excel file to the ZIP
            updated_excel_buffer = io.BytesIO()
            workbook.save(updated_excel_buffer)
            updated_excel_buffer.seek(0)
            zf.writestr("updated_file.xlsx", updated_excel_buffer.read())

        zip_buffer.seek(0)

        # Return ZIP file
        return send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name="etiquettes_with_excel.zip"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))  # Default to 5000 if PORT is not set
    app.run(host="0.0.0.0", port=port, debug=True)
